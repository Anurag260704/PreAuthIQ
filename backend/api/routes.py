"""Versioned API routes for MedClear."""

import json
import os
import re
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Union, cast

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from typing_extensions import TypedDict

from api.dependencies import get_benchmark_semaphore, get_engine, verify_api_key
from core.appeal import generate_appeal_draft
from core.config import (
    DEFAULT_MISTRAL_MODEL,
    MAX_RAW_CLINICAL_NOTES_LENGTH,
    MAX_UPLOAD_FILE_SIZE,
)
from core.models import AppealDraftResponse, PreAuthCaseInput, PreAuthSkillOutput
from core.response_parser import ComplexCaseRow, SchemaFieldRow, TrainingCaseRow, parse_complex_case, parse_training_cases
from core.training_notes import build_expanded_clinical_notes

router = APIRouter(prefix="/api/v1")
legacy_router = APIRouter(prefix="/api", include_in_schema=False)

DATA_DIR = Path(__file__).parent.parent / "data"


class ValidationResult(TypedDict):
    case_id: str
    requested_service: Optional[str]
    expected: str
    actual: str
    match: bool
    confidence: Optional[str]
    complexity_notes: Optional[str]
    note: Optional[str]


class ValidationSummary(TypedDict):
    total: int
    correct: int
    accuracy: float
    results: List[ValidationResult]


def _load_json(filename: str) -> object:
    path = DATA_DIR / filename
    if not path.exists():
        raise HTTPException(
            status_code=503,
            detail=f"Data file not found: {filename}. Run scripts/parse_workbook.py first.",
        )
    with open(path, encoding="utf-8") as file:
        return json.load(file)


def _load_training_cases() -> List[TrainingCaseRow]:
    return cast(List[TrainingCaseRow], _load_json("training_cases.json"))


def _load_schema_fields() -> List[SchemaFieldRow]:
    return cast(List[SchemaFieldRow], _load_json("patient_data_schema.json"))


def _complex_case_rows_to_input(rows: List[ComplexCaseRow]) -> PreAuthCaseInput:
    from skill.schema import SourcedField

    field_map: Dict[str, ComplexCaseRow] = {
        str(r.get("field", "")).lower().strip(): r for r in rows if r.get("field")
    }

    def _get_value(name: str) -> Optional[str]:
        row = field_map.get(name.lower())
        return row.get("value") if row else None

    def _get_sourced(name: str) -> Optional[SourcedField]:
        row = field_map.get(name.lower())
        if not row or not row.get("value"):
            return None
        return SourcedField(value=row.get("value"), source=row.get("source"), source_date=None)

    def _split_csv(value: Optional[str]) -> Optional[List[str]]:
        if not value:
            return None
        items = [item.strip() for item in str(value).replace(";", ",").split(",") if item.strip()]
        return items or None

    age_raw = _get_value("age / sex") or _get_value("age")
    age: Optional[int] = None
    sex: Optional[str] = None
    if age_raw:
        parts = [part.strip() for part in str(age_raw).split("/")]
        if parts and parts[0]:
            try:
                age = int(parts[0])
            except (ValueError, IndexError):
                age = None
        if len(parts) > 1 and parts[1]:
            sex = parts[1]

    requested_service_row = field_map.get("requested service") or {}
    requested_service_notes = str(requested_service_row.get("notes") or "")

    site_of_care = _get_value("site of care") or _get_value("site_of_care")
    if not site_of_care and "inpatient" in requested_service_notes.lower():
        site_of_care = "Inpatient"

    requested_los = _get_value("expected los")
    if not requested_los and "midnight" in requested_service_notes.lower():
        requested_los = "2 midnights"

    secondary_diagnoses = (
        _split_csv(_get_value("secondary diagnoses"))
        or _split_csv(_get_value("comorbidities"))
        or _split_csv(_get_value("medical risk factors"))
    )

    prior_conservative_treatment = (
        _split_csv(_get_value("conservative care"))
        or _split_csv(_get_value("prior conservative treatment"))
    )

    current_medications = _split_csv(_get_value("current medications"))

    conservative_response = _get_value("response")
    if not conservative_response and prior_conservative_treatment:
        conservative_response = "Conservative measures documented with incomplete symptom relief."

    return PreAuthCaseInput(
        case_id=_get_value("case_id") or "PA-001",
        payer_plan=_get_value("payer / plan") or _get_value("payer plan"),
        requested_service=_get_value("requested service") or "Unknown",
        site_of_care=site_of_care,
        requested_los=requested_los,
        payer_policy_version=_get_value("payer policy used") or _get_value("payer policy"),
        age=age,
        sex=sex,
        primary_diagnosis=_get_value("primary diagnosis") or "Unknown",
        secondary_diagnoses=secondary_diagnoses,
        symptom_duration=_get_value("symptom duration"),
        pain_severity=_get_value("pain severity"),
        functional_impairment_adls=_get_sourced("adl impact"),
        objective_neurologic_deficits=_get_sourced("objective findings"),
        imaging_findings=_get_sourced("mri cervical spine") or _get_sourced("imaging"),
        prior_conservative_treatment=prior_conservative_treatment,
        response_to_prior_treatment=conservative_response,
        current_medications=current_medications,
        complications_red_flags=_get_value("recent events"),
        contradictory_flags=_get_value("contradiction"),
        missing_records=_get_value("missing documentation"),
        ordering_provider_specialty=_get_value("ordering provider specialty"),
        raw_clinical_notes=_append_notes(
            None,
            "\n".join(
                [
                    f"Requested service notes: {requested_service_notes}",
                    f"Conservative care: {_get_value('conservative care') or 'not explicitly listed'}",
                    f"Response to treatment: {conservative_response or 'not specified'}",
                ]
            ),
        ),
    )


def _append_notes(existing: Optional[str], addition: str) -> Optional[str]:
    merged = addition.strip() if not existing else f"{existing.strip()}\n\n{addition.strip()}"
    merged = merged.strip()
    if not merged:
        return None
    return merged[:MAX_RAW_CLINICAL_NOTES_LENGTH]


def _serialize_workbook_context(workbook) -> str:
    max_rows = int(os.environ.get("MAX_WORKBOOK_ROWS_FOR_CONTEXT", "2500"))
    max_chars = min(
        int(os.environ.get("MAX_WORKBOOK_CONTEXT_CHARS", str(MAX_RAW_CLINICAL_NOTES_LENGTH))),
        MAX_RAW_CLINICAL_NOTES_LENGTH,
    )
    rows_seen = 0
    lines: List[str] = []

    for worksheet in workbook.worksheets:
        lines.append(f"[Sheet: {worksheet.title}]")
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [
                str(value).strip()
                for value in row
                if value is not None and str(value).strip()
            ]
            if not values:
                continue
            rows_seen += 1
            lines.append(f"Row {row_index}: " + " | ".join(values))
            context = "\n".join(lines)
            if rows_seen >= max_rows or len(context) >= max_chars:
                lines.append(
                    f"... workbook context truncated after {rows_seen} rows to fit processing limits."
                )
                return "\n".join(lines)[:max_chars]
    return "\n".join(lines)[:max_chars]


def _extract_label_value(text: str, labels: List[str]) -> Optional[str]:
    pattern = r"(?i)(" + "|".join(re.escape(label) for label in labels) + r")\s*[:\-]\s*(.+)"
    for line in text.splitlines():
        match = re.search(pattern, line)
        if match:
            value = match.group(2).strip()
            if value:
                return value[:500]
    return None


def _infer_requested_service(context: str) -> str:
    return _extract_label_value(
        context,
        [
            "requested service",
            "service requested",
            "procedure",
        ],
    ) or "Unknown requested service"


def _infer_primary_diagnosis(context: str) -> str:
    return _extract_label_value(
        context,
        [
            "primary diagnosis",
            "diagnosis",
            "clinical scenario",
        ],
    ) or "Unknown diagnosis"


@router.get("/status", tags=["System"], summary="API health check")
async def status() -> Dict[str, object]:
    return {"status": "ok", "model": DEFAULT_MISTRAL_MODEL}


@router.post(
    "/review",
    response_model=PreAuthSkillOutput,
    tags=["Review"],
    summary="Run full pipeline on structured case JSON",
)
async def review_case(
    clinical_input: PreAuthCaseInput,
    engine=Depends(get_engine),
) -> PreAuthSkillOutput:
    return await engine.process_case(clinical_input)


@router.post(
    "/review/appeal",
    response_model=AppealDraftResponse,
    tags=["Review"],
    summary="Generate appeal draft from a completed report",
)
async def review_appeal(report: PreAuthSkillOutput) -> AppealDraftResponse:
    return generate_appeal_draft(report)


@router.post(
    "/review/upload",
    response_model=PreAuthSkillOutput,
    tags=["Review"],
    summary="Upload assignment .xlsx workbook and run pipeline",
)
async def review_upload(
    file: UploadFile = File(...),
    engine=Depends(get_engine),
) -> PreAuthSkillOutput:
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    file_size = 0
    chunks: List[bytes] = []
    while True:
        chunk = await file.read(8192)
        if not chunk:
            break
        file_size += len(chunk)
        if file_size > MAX_UPLOAD_FILE_SIZE:
            raise HTTPException(status_code=413, detail="Uploaded file is too large.")
        chunks.append(chunk)
    content = b"".join(chunks)

    tmp_path: Optional[str] = None
    workbook = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
            tmp.write(content)
        workbook = load_workbook(tmp_path, read_only=True, data_only=True)
        workbook_context = _serialize_workbook_context(workbook)

        if "Complex_Case_Input" in workbook.sheetnames:
            case_input = _complex_case_rows_to_input(parse_complex_case(workbook))
            case_input.raw_clinical_notes = _append_notes(
                case_input.raw_clinical_notes,
                workbook_context,
            )
        elif "Training_Cases" in workbook.sheetnames:
            training_cases = parse_training_cases(workbook)
            if not training_cases:
                raise HTTPException(status_code=422, detail="Training_Cases sheet is empty.")
            first = training_cases[0]
            summary_notes = build_expanded_clinical_notes(first)
            case_input = PreAuthCaseInput(
                case_id=first.get("case_id") or "unknown",
                requested_service=first.get("requested_service") or _infer_requested_service(workbook_context),
                primary_diagnosis=first.get("clinical_scenario") or _infer_primary_diagnosis(workbook_context),
                raw_clinical_notes=_append_notes(summary_notes, workbook_context),
            )
        else:
            if not workbook_context.strip():
                raise HTTPException(
                    status_code=422,
                    detail="Workbook did not contain readable non-empty rows.",
                )
            case_input = PreAuthCaseInput(
                case_id="uploaded-workbook",
                requested_service=_infer_requested_service(workbook_context),
                primary_diagnosis=_infer_primary_diagnosis(workbook_context),
                raw_clinical_notes=workbook_context,
            )
        return await engine.process_case(case_input)
    finally:
        if workbook is not None:
            workbook.close()
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        await file.close()


@router.get("/samples", tags=["Samples"], summary="List all training cases")
async def list_samples() -> List[TrainingCaseRow]:
    return _load_training_cases()


@router.get(
    "/samples/complex",
    tags=["Samples"],
    summary="Get PA-001 complex-case input (assignment packet)",
)
async def get_complex_sample() -> PreAuthCaseInput:
    rows = _load_json("complex_case.json")
    return _complex_case_rows_to_input(cast(List[ComplexCaseRow], rows))


@router.get("/samples/{sample_id}", tags=["Samples"], summary="Get one training case by ID")
async def get_sample(sample_id: str) -> Dict[str, object]:
    for case in _load_training_cases():
        if case.get("case_id") == sample_id:
            result = dict(case)
            result["has_full_packet"] = sample_id == "PA-001"
            return result
    raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found.")


@router.get("/fields", tags=["Schema"], summary="List patient-data field glossary")
async def list_fields() -> List[SchemaFieldRow]:
    return _load_schema_fields()


@router.get("/fields/schema", tags=["Schema"], summary="OpenAPI JSON schemas for input and output")
async def get_fields_schema() -> Dict[str, object]:
    return {
        "input": PreAuthCaseInput.model_json_schema(),
        "output": PreAuthSkillOutput.model_json_schema(),
    }


@router.get(
    "/benchmark",
    tags=["Benchmark"],
    summary="Run accuracy benchmark on all training cases",
)
async def run_benchmark(_: str = Depends(verify_api_key), engine=Depends(get_engine)) -> ValidationSummary:
    semaphore = get_benchmark_semaphore()
    async with semaphore:
        cases = _load_training_cases()
        results: List[ValidationResult] = []
        for case in cases:
            case_id = case.get("case_id") or "unknown"
            expected = case.get("expected_outcome") or "NEED_MORE_INFO"
            input_case = PreAuthCaseInput(
                case_id=case_id,
                requested_service=case.get("requested_service") or "Unknown",
                primary_diagnosis=case.get("clinical_scenario") or "Unknown",
                raw_clinical_notes=build_expanded_clinical_notes(case)[:MAX_RAW_CLINICAL_NOTES_LENGTH],
            )
            try:
                output = await engine.process_case(input_case)
                results.append(
                    {
                        "case_id": case_id,
                        "requested_service": case.get("requested_service"),
                        "expected": expected,
                        "actual": output.recommendation,
                        "match": output.recommendation == expected,
                        "confidence": output.confidence,
                        "complexity_notes": case.get("complexity_notes"),
                        "note": None,
                    }
                )
            except Exception as exc:
                results.append(
                    {
                        "case_id": case_id,
                        "requested_service": case.get("requested_service"),
                        "expected": expected,
                        "actual": "ERROR",
                        "match": False,
                        "confidence": None,
                        "complexity_notes": case.get("complexity_notes"),
                        "note": str(exc),
                    }
                )
        correct = sum(1 for result in results if result["match"])
        return {
            "total": len(results),
            "correct": correct,
            "accuracy": round(correct / len(results), 2) if results else 0.0,
            "results": results,
        }


class ProgressEvent(TypedDict):
    type: str
    current: int
    total: int
    case_id: Optional[str]
    result: Optional[Union[ValidationResult, ValidationSummary]]


@router.get(
    "/benchmark/stream",
    tags=["Benchmark"],
    summary="Benchmark with Server-Sent Events progress stream",
)
async def run_benchmark_stream(
    _: str = Depends(verify_api_key),
    engine=Depends(get_engine),
) -> StreamingResponse:
    async def event_generator():
        summary = await run_benchmark("ok", engine)
        total = summary["total"]
        for index, result in enumerate(summary["results"]):
            progress: ProgressEvent = {
                "type": "progress",
                "current": index + 1,
                "total": total,
                "case_id": result["case_id"],
                "result": result,
            }
            yield f"data: {json.dumps(progress)}\n\n"
        done: ProgressEvent = {
            "type": "complete",
            "current": total,
            "total": total,
            "case_id": None,
            "result": summary,
        }
        yield f"data: {json.dumps(done)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"},
    )


def _legacy_aliases() -> None:
    """Backward-compatible /api/* paths (hidden from Swagger — use /api/v1/*)."""
    legacy_router.add_api_route("/health", status, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route(
        "/analyze", review_case, methods=["POST"], response_model=PreAuthSkillOutput, include_in_schema=False
    )
    legacy_router.add_api_route(
        "/analyze/appeal",
        review_appeal,
        methods=["POST"],
        response_model=AppealDraftResponse,
        include_in_schema=False,
    )
    legacy_router.add_api_route(
        "/analyze/upload",
        review_upload,
        methods=["POST"],
        response_model=PreAuthSkillOutput,
        include_in_schema=False,
    )
    legacy_router.add_api_route("/cases", list_samples, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/cases/{sample_id}", get_sample, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/schema", list_fields, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/schema/json", get_fields_schema, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/complex-case/input", get_complex_sample, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/validate-all", run_benchmark, methods=["GET"], include_in_schema=False)
    legacy_router.add_api_route("/validate-all/stream", run_benchmark_stream, methods=["GET"], include_in_schema=False)


_legacy_aliases()
